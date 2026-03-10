import subprocess
import sys

# Instalar dependencias
subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "pandas"], capture_output=True)

import pandas as pd
import openpyxl

archivo = r"C:\Users\aaron\OneDrive\Escritorio\Registro UTN\Automatizarcm\Listado de personas matriculadas 2026.xlsx"

# ---- 1. Hojas del libro ----
wb = openpyxl.load_workbook(archivo, data_only=True)
print("=" * 70)
print(f"HOJAS EN EL ARCHIVO: {wb.sheetnames}")
print("=" * 70)

for hoja in wb.sheetnames:
    print(f"\n{'=' * 70}")
    print(f"HOJA: '{hoja}'")
    print("=" * 70)

    df = pd.read_excel(archivo, sheet_name=hoja, header=None)
    print(f"Dimensiones brutas: {df.shape[0]} filas x {df.shape[1]} columnas")

    # Intentar detectar fila de encabezado (primera fila no vacía)
    for i, row in df.iterrows():
        if row.notna().sum() > 2:
            header_row = i
            break

    df_clean = pd.read_excel(archivo, sheet_name=hoja, header=header_row)
    df_clean.columns = [str(c).strip() for c in df_clean.columns]

    print(f"\nColumnas ({len(df_clean.columns)}):")
    for j, col in enumerate(df_clean.columns):
        dtype = df_clean[col].dtype
        no_nulos = df_clean[col].notna().sum()
        print(f"  [{j+1:02d}] '{col}'  |  tipo: {dtype}  |  valores no nulos: {no_nulos}")

    print(f"\nTotal de filas de datos: {len(df_clean)}")

    # Muestra de valores únicos por columna (máx 10)
    print("\nMuestra de valores únicos por columna:")
    for col in df_clean.columns:
        unicos = df_clean[col].dropna().unique()
        muestra = list(unicos[:10])
        print(f"  '{col}': {muestra}")

    # Primeras 5 filas
    print("\nPrimeras 5 filas:")
    print(df_clean.head(5).to_string(index=False))

print("\n" + "=" * 70)
print("ANÁLISIS COMPLETO FINALIZADO")
print("=" * 70)
